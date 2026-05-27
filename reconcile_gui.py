import os
import re
import datetime
import threading
from decimal import Decimal, InvalidOperation

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

import openpyxl

# ==================== 业务配置常量 ====================
BANK_KEYWORDS = ["银行", "流水", "对账单"]
BANK_HEADER_ROW = 16
BANK_DATA_START_ROW = 17

PR_KEYWORDS = ["预收款"]
PR_HEADER_ROW = 3
PR_DATA_START_ROW = 4

# 预收款表格对应的精确字段名称
COL_PR_AMOUNT = "*收款金额 # amount"
COL_PR_BILLNO = "*单据编号 # billno"
COL_PR_CLIENT = "*结算客户.名称 # itemclass.name"
COL_PR_ORDER = "*订单编号 # srcbillno"

# 银行流水对应的目标字段名称
COL_BANK_AMOUNT = "贷方发生额"

# 写入目标列 (1-based index, O=15, P=16, Q=17)
COL_O_INDEX = 15  # 单据编号
COL_P_INDEX = 16  # 客户
COL_Q_INDEX = 17  # 订单号
# ====================================================

class ReconciliationApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("银行流水与预收款自动核对工具")
        self.geometry("750x620")
        self.minsize(650, 520)
        
        # 绑定的变量
        self.bank_file_path = tk.StringVar()
        self.pr_file_path = tk.StringVar()
        self.status_text = tk.StringVar(value="准备就绪")
        
        self.init_ui()
        self.auto_detect_files()

    def init_ui(self):
        # 1. 文件选择区域
        file_frame = ttk.LabelFrame(self, text="数据源选择", padding=10)
        file_frame.pack(fill=tk.X, padx=15, pady=10)

        # 银行流水文件
        ttk.Label(file_frame, text="银行流水文件:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.bank_entry = ttk.Entry(file_frame, textvariable=self.bank_file_path, width=55)
        self.bank_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(file_frame, text="选择...", command=self.browse_bank_file).grid(row=0, column=2, padx=5, pady=5)

        # 预收款文件
        ttk.Label(file_frame, text="预收款文件:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.pr_entry = ttk.Entry(file_frame, textvariable=self.pr_file_path, width=55)
        self.pr_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(file_frame, text="选择...", command=self.browse_pr_file).grid(row=1, column=2, padx=5, pady=5)

        file_frame.columnconfigure(1, weight=1)
        
        btn_frame = ttk.Frame(file_frame)
        btn_frame.grid(row=2, column=0, columnspan=3, pady=5, sticky=tk.E)
        ttk.Button(btn_frame, text="重新自动识别当前目录", command=self.auto_detect_files).pack(side=tk.RIGHT)

        # 2. 控制与状态区域
        control_frame = ttk.Frame(self, padding=5)
        control_frame.pack(fill=tk.X, padx=15, pady=5)

        self.run_btn = ttk.Button(control_frame, text="开始执行对账核对", command=self.start_processing_thread)
        self.run_btn.pack(side=tk.LEFT, ipadx=15)

        self.status_label = ttk.Label(control_frame, textvariable=self.status_text, font=("Arial", 10, "bold"), foreground="blue")
        self.status_label.pack(side=tk.RIGHT, padx=10)

        # 3. 日志输出区域
        log_frame = ttk.LabelFrame(self, text="处理日志输出", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        self.log_area = ScrolledText(log_frame, wrap=tk.WORD, height=15)
        self.log_area.pack(fill=tk.BOTH, expand=True)
        self.log_area.configure(state='disabled')

    def log(self, message):
        self.log_area.configure(state='normal')
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.log_area.configure(state='disabled')

    def clear_log(self):
        self.log_area.configure(state='normal')
        self.log_area.delete(1.0, tk.END)
        self.log_area.configure(state='disabled')

    def auto_detect_files(self):
        current_dir = os.getcwd()
        bank_file = self.find_file_by_keywords(current_dir, BANK_KEYWORDS)
        pr_file = self.find_file_by_keywords(current_dir, PR_KEYWORDS)
        
        if bank_file:
            self.bank_file_path.set(bank_file)
            self.log(f"[系统自动识别] 已定位到银行流水文件: {os.path.basename(bank_file)}")
        else:
            self.bank_file_path.set("")
            self.log("[提示] 未在当前目录自动识别到银行流水文件，请手动选择。")

        if pr_file:
            self.pr_file_path.set(pr_file)
            self.log(f"[系统自动识别] 已定位到预收款文件: {os.path.basename(pr_file)}")
        else:
            self.pr_file_path.set("")
            self.log("[提示] 未在当前目录自动识别到预收款文件，请手动选择。")

    def find_file_by_keywords(self, directory, keywords):
        for filename in os.listdir(directory):
            if not filename.endswith('.xlsx') or filename.startswith('~$'):
                continue
            if any(kw in filename for kw in keywords):
                return os.path.join(directory, filename)
        return None

    def browse_bank_file(self):
        file_selected = filedialog.askopenfilename(
            title="选择银行流水Excel文件",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if file_selected:
            normalized_path = os.path.normpath(file_selected)
            self.bank_file_path.set(normalized_path)
            self.log(f"[手动选择] 银行流水文件已更改为: {os.path.basename(normalized_path)}")

    def browse_pr_file(self):
        file_selected = filedialog.askopenfilename(
            title="选择预收款Excel文件",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if file_selected:
            normalized_path = os.path.normpath(file_selected)
            self.pr_file_path.set(normalized_path)
            self.log(f"[手动选择] 预收款文件已更改为: {os.path.basename(normalized_path)}")

    def clean_amount(self, val):
        if val is None:
            return None
        val_str = str(val).strip().replace(',', '')
        if not val_str:
            return None
        try:
            cleaned = re.sub(r'[^\d\.\-]', '', val_str)
            return Decimal(cleaned)
        except (InvalidOperation, ValueError):
            return None

    def get_sheet_insensitive(self, workbook, target_name="sheet1"):
        for name in workbook.sheetnames:
            if name.lower() == target_name.lower():
                return workbook[name]
        return workbook.active

    def start_processing_thread(self):
        if not self.bank_file_path.get() or not self.pr_file_path.get():
            messagebox.showwarning("警告", "请确保已指定或选择了两类数据文件！")
            return
        
        self.run_btn.configure(state=tk.DISABLED)
        self.status_text.set("正在核对数据中...")
        self.clear_log()
        
        t = threading.Thread(target=self.process_reconciliation)
        t.daemon = True
        t.start()

    def process_reconciliation(self):
        start_time = datetime.datetime.now()
        log_messages = []

        def local_log(msg):
            self.log(msg)
            log_messages.append(msg)

        bank_path = self.bank_file_path.get()
        pr_path = self.pr_file_path.get()

        target_output_dir = os.path.dirname(bank_path)

        local_log(f"开始处理时间: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        local_log(f"银行流水路径: {bank_path}")
        local_log(f"预收款路径: {pr_path}")
        local_log(f"保存目标目录: {target_output_dir}")

        # 1. 读取预收款文件
        try:
            pr_wb = openpyxl.load_workbook(pr_path, data_only=True)
            pr_ws = self.get_sheet_insensitive(pr_wb, "sheet1")
        except Exception as e:
            local_log(f"错误: 读取预收款文件失败: {str(e)}")
            self.finalize_run(False, log_messages, target_output_dir)
            return

        pr_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in pr_ws[PR_HEADER_ROW]]
        
        pr_col_amount_idx = None
        pr_col_billno_idx = None
        pr_col_client_idx = None
        pr_col_order_idx = None

        # 精准匹配字段表头
        for idx, header in enumerate(pr_headers, start=1):
            h_clean = header.strip()
            if h_clean == COL_PR_AMOUNT  in h_clean.lower():
                pr_col_amount_idx = idx
            if h_clean == COL_PR_BILLNO  in h_clean.lower():
                pr_col_billno_idx = idx
            if h_clean == COL_PR_CLIENT or "itemclass.name" in h_clean.lower():
                pr_col_client_idx = idx
            if h_clean == COL_PR_ORDER or "srcbillno" in h_clean.lower():
                pr_col_order_idx = idx

        missing_pr_cols = []
        if not pr_col_amount_idx: missing_pr_cols.append(COL_PR_AMOUNT)
        if not pr_col_billno_idx: missing_pr_cols.append(COL_PR_BILLNO)
        if not pr_col_client_idx: missing_pr_cols.append(COL_PR_CLIENT)
        if not pr_col_order_idx: missing_pr_cols.append(COL_PR_ORDER)

        if missing_pr_cols:
            local_log(f"错误：预收款文件中缺少必要列: {', '.join(missing_pr_cols)}")
            self.finalize_run(False, log_messages, target_output_dir)
            return

        pr_data = {}
        total_pr_rows = 0

        for r in range(PR_DATA_START_ROW, pr_ws.max_row + 1):
            amount_val = pr_ws.cell(row=r, column=pr_col_amount_idx).value
            billno_val = pr_ws.cell(row=r, column=pr_col_billno_idx).value
            client_val = pr_ws.cell(row=r, column=pr_col_client_idx).value
            order_val = pr_ws.cell(row=r, column=pr_col_order_idx).value

            if amount_val is None and billno_val is None and client_val is None and order_val is None:
                continue

            total_pr_rows += 1
            amount = self.clean_amount(amount_val)
            if amount is None:
                local_log(f"提示 [预收款行号 {r}]: 金额为空或无法解析。")
                continue

            item = {
                "billno": str(billno_val).strip() if billno_val is not None else "",
                "client": str(client_val).strip() if client_val is not None else "",
                "order": str(order_val).strip() if order_val is not None else "",
                "row_num": r
            }
            pr_data.setdefault(amount, []).append(item)

        local_log(f"预收款有效数据记录数: {total_pr_rows}")

        # 2. 读取银行流水文件
        try:
            bank_wb_values = openpyxl.load_workbook(bank_path, data_only=True)
            bank_ws_values = self.get_sheet_insensitive(bank_wb_values, "sheet1")
            
            bank_wb_save = openpyxl.load_workbook(bank_path, data_only=False)
            bank_ws_save = self.get_sheet_insensitive(bank_wb_save, "sheet1")
        except Exception as e:
            local_log(f"错误: 读取银行流水文件失败: {str(e)}")
            self.finalize_run(False, log_messages, target_output_dir)
            return

        bank_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in bank_ws_values[BANK_HEADER_ROW]]
        bank_col_amount_idx = None
        for idx, header in enumerate(bank_headers, start=1):
            if COL_BANK_AMOUNT in header:
                bank_col_amount_idx = idx
                break

        if not bank_col_amount_idx:
            local_log(f"错误：银行流水文件中未找到「{COL_BANK_AMOUNT}」列。")
            self.finalize_run(False, log_messages, target_output_dir)
            return

        bank_rows = {}
        total_bank_rows = 0

        for r in range(BANK_DATA_START_ROW, bank_ws_values.max_row + 1):
            amount_val = bank_ws_values.cell(row=r, column=bank_col_amount_idx).value
            row_vals = [bank_ws_values.cell(row=r, column=c).value for c in range(1, bank_ws_values.max_column + 1)]
            if all(v is None for v in row_vals):
                continue

            total_bank_rows += 1
            amount = self.clean_amount(amount_val)
            if amount is None:
                local_log(f"提示 [银行流水行号 {r}]: 金额为空或无法解析。")
                continue

            bank_rows.setdefault(amount, []).append(r)

        local_log(f"银行流水有效数据记录数: {total_bank_rows}")

        # 3. 对账比对与填充
        matched_count = 0
        unmatched_count = 0
        
        bank_ws_save.cell(row=BANK_HEADER_ROW, column=COL_O_INDEX, value="单据编号")
        bank_ws_save.cell(row=BANK_HEADER_ROW, column=COL_P_INDEX, value="客户")
        bank_ws_save.cell(row=BANK_HEADER_ROW, column=COL_Q_INDEX, value="订单号")

        all_amounts = set(bank_rows.keys()).union(set(pr_data.keys()))

        for amount in all_amounts:
            b_rows = bank_rows.get(amount, [])
            p_items = pr_data.get(amount, [])

            len_b = len(b_rows)
            len_p = len(p_items)

            if len_b == 0:
                continue

            if len_b == 1 and len_p == 1:
                bank_row = b_rows[0]
                pr_item = p_items[0]

                # 将预收款的 "*单据编号 # billno" 数据拷贝至 O 列
                bank_ws_save.cell(row=bank_row, column=COL_O_INDEX, value=pr_item["billno"])
                # 将预收款的 "*结算客户.名称 # itemclass.name" 数据拷贝至 P 列
                bank_ws_save.cell(row=bank_row, column=COL_P_INDEX, value=pr_item["client"])
                # 将预收款的 "*订单编号 # srcbillno" 数据拷贝至 Q 列
                bank_ws_save.cell(row=bank_row, column=COL_Q_INDEX, value=pr_item["order"])
                
                empty_fields = []
                if not pr_item["billno"]: empty_fields.append(COL_PR_BILLNO)
                if not pr_item["client"]: empty_fields.append(COL_PR_CLIENT)
                if not pr_item["order"]: empty_fields.append(COL_PR_ORDER)
                if empty_fields:
                    local_log(f"提示 [信息缺失]: 金额 {amount} 在流水行 {bank_row} 匹配成功，但预收款行 {pr_item['row_num']} 对应的字段 {', '.join(empty_fields)} 为空值。")

                matched_count += 1
            elif len_b > 1 or len_p > 1:
                unmatched_count += len_b
                local_log(f"无法匹配 [金额重复]: 金额 {amount} 在流水中出现 {len_b} 次，在预收款中出现 {len_p} 次，无法确保精确对应，已跳过。")
            else:
                unmatched_count += len_b

        # 4. 组装输出路径
        end_time = datetime.datetime.now()
        timestamp_str = end_time.strftime("%Y%m%d_%H%M%S")
        
        output_excel_path = os.path.join(target_output_dir, f"{timestamp_str}.xlsx")
        output_log_path = os.path.join(target_output_dir, f"{timestamp_str}.txt")
        
        # 5. 保存结果文件
        try:
            bank_wb_save.save(output_excel_path)
            local_log(f"结果Excel文件已成功保存至: {output_excel_path}")
        except Exception as e:
            local_log(f"错误: 保存结果文件失败: {str(e)}")
            self.finalize_run(False, log_messages, target_output_dir)
            return

        local_log("\n" + "="*15 + " 核对统计汇总 " + "="*15)
        local_log(f"完成处理时间: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        local_log(f"数据比对耗时: {(end_time - start_time).total_seconds():.2f} 秒")
        local_log(f"银行流水有效总记录: {total_bank_rows}")
        local_log(f"成功匹配关联记录: {matched_count}")
        local_log(f"未成功匹配记录: {unmatched_count}")
        local_log("="*44)

        # 6. 保存TXT日志文件
        try:
            with open(output_log_path, "w", encoding="utf-8") as f:
                for line in log_messages:
                    f.write(line + "\n")
            local_log(f"日志文件已成功保存至: {output_log_path}")
        except Exception as e:
            local_log(f"警告: 无法保存日志文件: {str(e)}")

        self.finalize_run(True, log_messages, target_output_dir)

    def finalize_run(self, success, log_messages, target_output_dir):
        self.run_btn.configure(state=tk.NORMAL)
        if success:
            self.status_text.set("核对完毕")
            messagebox.showinfo("成功", f"账目数据比对执行完成！\n\n核对后的Excel文件和TXT日志已全部保存至以下文件夹目录：\n\n{target_output_dir}")
        else:
            self.status_text.set("核对失败")
            messagebox.showerror("出错", "在对账过程中遇到了错误，请检查日志。")


if __name__ == "__main__":
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = ReconciliationApp()
    app.mainloop()