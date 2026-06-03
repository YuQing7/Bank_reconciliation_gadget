import os
import re
import csv
import datetime
import threading
from decimal import Decimal, InvalidOperation

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

import openpyxl

# ==================== 业务配置常量 ====================
PR_KEYWORDS = ["预收款"]
PR_HEADER_ROW = 1
PR_DATA_START_ROW = 2

# 三种模式的匹配配置
MODE_CONFIGS = {
    "银行": {
        "bank_keywords": ["银行", "流水", "对账单"],
        "bank_header_row": 16,
        "bank_data_start_row": 17,
        "pr_settle_account": "中信银行5031",
        "bank_cols": {
            "amount": "贷方发生额",
            "date": "交易日期",
            "opp_name": "对方账户名称",
            "currency": "币种"
        },
        "output_cols": {
            15: ("单据编号", "billno"),        # O 列
            16: ("结算客户编码", "client_code"), # P 列
            17: ("结算客户", "client"),       # Q 列
            18: ("收款金额", "amount"),        # R 列
            19: ("收款金额本位币", "amount_local") # S 列
        }
    },
    "支付宝": {
        "bank_keywords": ["支付宝", "alipay"],
        "bank_header_row": 3,
        "bank_data_start_row": 4,
        "pr_settle_account": "支付宝",
        "bank_cols": {
            "amount": "收入（+元）",
            "date": "入账时间"
        },
        "output_cols": {
            23: ("单据编号", "billno"),        # W 列
            24: ("结算客户编码", "client_code"), # X 列
            25: ("结算客户", "client"),       # Y 列
            26: ("收款金额本位币", "amount_local"), # Z 列
            27: ("收款备注", "comment")        # AA 列
        }
    },
    "微信": {
        "bank_keywords": ["微信", "wechat"],
        "bank_header_row": 1,
        "bank_data_start_row": 2,
        "pr_settle_account": "宇问测量微信",
        "bank_cols": {
            "amount": "实收金额（元）",
            "date": "交易时间"
        },
        "output_cols": {
            39: ("单据编号", "billno"),        # AM 列
            40: ("结算客户编码", "client_code"), # AN 列
            41: ("结算客户", "client"),       # AO 列
            42: ("收款金额本位币", "amount_local"), # AP 列
            43: ("收款备注", "comment")        # AQ 列
        }
    }
}
# ====================================================

class ReconciliationApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("流水自动核对工具")
        self.geometry("750x660")
        self.minsize(650, 550)
        
        # 绑定的变量
        self.mode_var = tk.StringVar(value="银行")
        self.bank_file_path = tk.StringVar()
        self.pr_file_path = tk.StringVar()
        self.status_text = tk.StringVar(value="准备就绪")
        
        self.init_ui()
        self.auto_detect_files()

    def init_ui(self):
        # 1. 模式选择区域
        mode_frame = ttk.LabelFrame(self, text="请选择对账模式", padding=10)
        mode_frame.pack(fill=tk.X, padx=15, pady=8)
        
        ttk.Radiobutton(mode_frame, text=" 银行 ", variable=self.mode_var, value="银行", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)
        ttk.Radiobutton(mode_frame, text=" 支付宝 ", variable=self.mode_var, value="支付宝", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)
        ttk.Radiobutton(mode_frame, text=" 微信 ", variable=self.mode_var, value="微信", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)

        # 2. 数据源选择区域
        file_frame = ttk.LabelFrame(self, text="数据源选择", padding=10)
        file_frame.pack(fill=tk.X, padx=15, pady=5)

        # 流水文件
        self.bank_label = ttk.Label(file_frame, text="流水文件:")
        self.bank_label.grid(row=0, column=0, sticky=tk.W, pady=5)
        self.bank_entry = ttk.Entry(file_frame, textvariable=self.bank_file_path, width=55)
        self.bank_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(file_frame, text="选择...", command=self.browse_bank_file).grid(row=0, column=2, padx=5, pady=5)

        # 预收款文件
        ttk.Label(file_frame, text="预收款文件:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.pr_entry = ttk.Entry(file_frame, textvariable=self.pr_file_path, width=55)
        self.pr_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)
        ttk.Button(file_frame, text="选择...", command=self.browse_pr_file).grid(row=1, column=2, padx=5, pady=5)

        file_frame.columnconfigure(1, weight=1)

        # 3. 控制与状态区域
        control_frame = ttk.Frame(self, padding=5)
        control_frame.pack(fill=tk.X, padx=15, pady=5)

        self.run_btn = ttk.Button(control_frame, text="开始执行对账核对", command=self.start_processing_thread)
        self.run_btn.pack(side=tk.LEFT, ipadx=15)

        # 新增的软件说明按钮
        self.intro_btn = ttk.Button(control_frame, text="软件使用说明", command=self.show_software_intro)
        self.intro_btn.pack(side=tk.LEFT, padx=10, ipadx=10)

        self.status_label = ttk.Label(control_frame, textvariable=self.status_text, font=("Arial", 10, "bold"), foreground="blue")
        self.status_label.pack(side=tk.RIGHT, padx=10)

        # 4. 日志输出区域
        log_frame = ttk.LabelFrame(self, text="处理日志输出", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        self.log_area = ScrolledText(log_frame, wrap=tk.WORD, height=13)
        self.log_area.pack(fill=tk.BOTH, expand=True)
        self.log_area.configure(state='disabled')

    def log(self, message):
        self.log_area.configure(state='normal')
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.log_area.configure(state='disabled')

    def clear_log(self):
        self.log_area.configure(state='normal')
        self.log_area.delete(1.0, tk.END)
        self.log_area.configure(state='disabled')

    def on_mode_change(self):
        mode = self.mode_var.get()
        self.bank_label.config(text=f"{mode}流水文件:")
        self.log(f"[模式切换] 当前对账模式切换为: 【{mode}】")
        self.auto_detect_files()

    def auto_detect_files(self):
        current_dir = os.getcwd()
        mode = self.mode_var.get()
        cfg = MODE_CONFIGS[mode]

        bank_file = self.find_file_by_keywords(current_dir, cfg["bank_keywords"])
        pr_file = self.find_file_by_keywords(current_dir, PR_KEYWORDS)
        
        if bank_file:
            self.bank_file_path.set(bank_file)
            self.log(f"[系统自动识别] 已定位到【{mode}】流水文件: {os.path.basename(bank_file)}")
        else:
            self.bank_file_path.set("")
            self.log(f"[提示] 请手动选择【{mode}】流水文件")

        if pr_file:
            self.pr_file_path.set(pr_file)
            self.log(f"[系统自动识别] 已定位到预收款文件: {os.path.basename(pr_file)}")
        else:
            self.pr_file_path.set("")
            self.log("[提示] 请手动选择预收款文件")

    def find_file_by_keywords(self, directory, keywords):
        for filename in os.listdir(directory):
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ('.xlsx', '.xls', '.csv') or filename.startswith('~$'):
                continue
            if any(kw in filename for kw in keywords):
                return os.path.join(directory, filename)
        return None

    def browse_bank_file(self):
        mode = self.mode_var.get()
        file_selected = filedialog.askopenfilename(
            title=f"选择{mode}流水文件",
            filetypes=[("支持的文件", "*.xlsx;*.xls;*.csv"), ("Excel 文件", "*.xlsx;*.xls"), ("CSV 文件", "*.csv"), ("所有文件", "*.*")]
        )
        if file_selected:
            normalized_path = os.path.normpath(file_selected)
            self.bank_file_path.set(normalized_path)
            self.log(f"[手动选择] {mode}流水文件已更改为: {os.path.basename(normalized_path)}")

    def browse_pr_file(self):
        file_selected = filedialog.askopenfilename(
            title="选择预收款Excel文件",
            filetypes=[("支持的文件", "*.xlsx;*.xls;*.csv"), ("Excel 文件", "*.xlsx;*.xls"), ("CSV 文件", "*.csv"), ("所有文件", "*.*")]
        )
        if file_selected:
            normalized_path = os.path.normpath(file_selected)
            self.pr_file_path.set(normalized_path)
            self.log(f"[手动选择] 预收款文件已更改为: {os.path.basename(normalized_path)}")

    def clean_amount(self, val):
        if val is None:
            return None
        val_str = str(val).strip().replace(',', '')
        if not val_str:
            return None
        try:
            cleaned = re.sub(r'[^\d\.\-]', '', val_str)
            if not cleaned:
                return None
            return Decimal(cleaned)
        except (InvalidOperation, ValueError):
            return None

    def parse_date_only(self, val):
        if val is None:
            return None
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        val_str = str(val).strip()
        if not val_str:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
            try:
                return datetime.datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue
        match = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', val_str)
        if match:
            try:
                return datetime.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            except ValueError:
                pass
        return None

    def is_chinese_name(self, name):
        if not name:
            return False
        return bool(re.search(r'[\u4e00-\u9fa5]', str(name)))

    def clean_string_for_compare(self, s):
        """清洗文本：去除非字母、数字以及中文字符，并将英文字符统一转小写进行无差别比较"""
        if s is None:
            return ""
        s_str = str(s).strip()
        cleaned = re.sub(r'[^a-zA-Z0-9\u4e00-\u9fa5]', '', s_str)
        return cleaned.lower()

    def find_header_col(self, headers, targets):
        # 第一阶段：完全匹配优先（去除空格，不区分大小写）
        for idx, h in enumerate(headers, start=1):
            h_clean = h.strip().lower()
            for t in targets:
                if h_clean == t.strip().lower():
                    return idx
                    
        # 第二阶段：模糊匹配（仅在未找到完全匹配列时作为备选方案）
        for idx, h in enumerate(headers, start=1):
            h_clean = h.strip().lower()
            for t in targets:
                t_clean = t.strip().lower()
                if t_clean in h_clean:
                    # 特殊防御：如果寻找的目标是 "结算客户"，需防止其误匹配到含有 "编码" 的列
                    if t_clean == "结算客户" and "编码" in h_clean:
                        continue
                    return idx
        return None

    def get_sheet_insensitive(self, workbook, target_name="sheet1"):
        for name in workbook.sheetnames:
            if name.lower() == target_name.lower():
                return workbook[name]
        return workbook.active

    def load_any_file_as_openpyxl(self, file_path, data_only=True):
        """支持加载 .xlsx、.xls、.csv，并统一转换为 openpyxl.Workbook 返回"""
        ext = os.path.splitext(file_path)[1].lower()
        
        # 1. 处理 .xlsx 文件
        if ext == '.xlsx':
            return openpyxl.load_workbook(file_path, data_only=data_only)
            
        # 2. 处理 .xls 文件
        elif ext == '.xls':
            # 在执行传统读取前，先检测该 .xls 文件底层是否实际上是 XML 格式
            is_xml = False
            try:
                with open(file_path, 'rb') as f:
                    header_bytes = f.read(150)
                    if b'<?xml' in header_bytes:
                        is_xml = True
            except Exception:
                pass
                
            if is_xml:
                try:
                    return self.parse_excel_xml(file_path)
                except Exception as e:
                    raise ValueError(f"检测到此 .xls 文件是 XML 格式，但在尝试解析时出错: {str(e)}")
            
            # 确认是真正的二进制旧版 .xls，调用 xlrd 读取
            try:
                import xlrd
            except ImportError:
                raise ImportError("检测到您选择了旧版 .xls 格式文件，需安装 xlrd 库才能正常读取。请在终端执行 'pip install xlrd' 后重试。")
            
            xls_wb = xlrd.open_workbook(file_path)
            op_wb = openpyxl.Workbook()
            if op_wb.sheetnames:
                op_wb.remove(op_wb.active)
                
            for sheet_name in xls_wb.sheet_names():
                xls_sheet = xls_wb.sheet_by_name(sheet_name)
                op_sheet = op_wb.create_sheet(title=sheet_name)
                
                # 循环转换所有单元格
                for r in range(xls_sheet.nrows):
                    for c in range(xls_sheet.ncols):
                        cell_val = xls_sheet.cell_value(r, c)
                        cell_type = xls_sheet.cell_type(r, c)
                        if cell_type == xlrd.XL_CELL_DATE:
                            try:
                                dt_tuple = xlrd.xldate_as_tuple(cell_val, xls_wb.datemode)
                                cell_val = datetime.datetime(*dt_tuple)
                            except Exception:
                                pass
                        op_sheet.cell(row=r+1, column=c+1, value=cell_val)
                
                # 同步合并单元格属性
                for crange in xls_sheet.merged_cells:
                    rlo, rhi, clo, chi = crange
                    op_sheet.merge_cells(
                        start_row=rlo+1, end_row=rhi,
                        start_column=clo+1, end_column=chi
                    )
            return op_wb
            
        # 3. 处理 .csv 文件
        elif ext == '.csv':
            op_wb = openpyxl.Workbook()
            if op_wb.sheetnames:
                op_wb.remove(op_wb.active)
            op_sheet = op_wb.create_sheet(title="Sheet1")
            
            # 中文环境下常见的几种CSV字符编码集
            encodings = ['utf-8-sig', 'gbk', 'utf-8', 'gb18030', 'latin-1']
            csv_data = None
            for enc in encodings:
                try:
                    with open(file_path, mode='r', encoding=enc) as f:
                        reader = csv.reader(f)
                        csv_data = list(reader)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            
            if csv_data is None:
                raise UnicodeDecodeError(f"未能以支持的中文或英文编码方式正确读取此 CSV 文件，请检查该文件编码: {file_path}")
            
            for r_idx, row in enumerate(csv_data, start=1):
                for c_idx, val in enumerate(row, start=1):
                    op_sheet.cell(row=r_idx, column=c_idx, value=val)
            return op_wb
            
        else:
            raise ValueError(f"系统不接受该类型的文件格式: {ext}")
   
    def parse_excel_xml(self, file_path):
        """解析 XML 2003 格式的 '伪xls' 文件，并将其转换为 openpyxl.Workbook 对象"""
        import xml.etree.ElementTree as ET
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        op_wb = openpyxl.Workbook()
        if op_wb.sheetnames:
            op_wb.remove(op_wb.active)
            
        # 查找 Worksheet 节点
        worksheets = [elem for elem in root.iter() if elem.tag.split('}')[-1] == 'Worksheet']
        
        for ws_elem in worksheets:
            # 获取工作表名称
            ws_name = "Sheet"
            for attr_name, attr_val in ws_elem.attrib.items():
                if attr_name.split('}')[-1] == 'Name':
                    ws_name = attr_val
                    break
            
            op_sheet = op_wb.create_sheet(title=ws_name)
            
            # 查找 Table 节点并提取 Rows
            tables = [child for child in ws_elem if child.tag.split('}')[-1] == 'Table']
            rows = []
            if tables:
                for table in tables:
                    rows.extend([child for child in table if child.tag.split('}')[-1] == 'Row'])
            else:
                rows = [child for child in ws_elem if child.tag.split('}')[-1] == 'Row']
                
            r_idx = 1
            for row_elem in rows:
                # 处理可能跨行跳行的情况 (获取 ss:Index)
                row_idx_attr = None
                for attr_name, attr_val in row_elem.attrib.items():
                    if attr_name.split('}')[-1] == 'Index':
                        row_idx_attr = attr_val
                        break
                if row_idx_attr:
                    r_idx = int(row_idx_attr)
                    
                cells = [child for child in row_elem if child.tag.split('}')[-1] == 'Cell']
                
                c_idx = 1
                for cell_elem in cells:
                    # 处理可能跨列跳列的情况 (获取 ss:Index)
                    cell_idx_attr = None
                    for attr_name, attr_val in cell_elem.attrib.items():
                        if attr_name.split('}')[-1] == 'Index':
                            cell_idx_attr = attr_val
                            break
                    if cell_idx_attr:
                        c_idx = int(cell_idx_attr)
                        
                    # 寻找 Cell 下的 Data 节点
                    data_elem = None
                    for child in cell_elem:
                        if child.tag.split('}')[-1] == 'Data':
                            data_elem = child
                            break
                            
                    if data_elem is not None:
                        val = data_elem.text
                        val_type = None
                        for attr_name, attr_val in data_elem.attrib.items():
                            if attr_name.split('}')[-1] == 'Type':
                                val_type = attr_val
                                break
                                
                        if val is not None:
                            # 还原数字类型
                            if val_type == 'Number':
                                try:
                                    if '.' in val:
                                        val = float(val)
                                    else:
                                        val = int(val)
                                except ValueError:
                                    pass
                        op_sheet.cell(row=r_idx, column=c_idx, value=val)
                        
                    # 处理合并单元格属性 (ss:MergeAcross, ss:MergeDown)
                    merge_across = 0
                    merge_down = 0
                    for attr_name, attr_val in cell_elem.attrib.items():
                        local_attr = attr_name.split('}')[-1]
                        if local_attr == 'MergeAcross':
                            merge_across = int(attr_val)
                        elif local_attr == 'MergeDown':
                            merge_down = int(attr_val)
                    
                    if merge_across > 0 or merge_down > 0:
                        op_sheet.merge_cells(
                            start_row=r_idx, end_row=r_idx + merge_down,
                            start_column=c_idx, end_column=c_idx + merge_across
                        )
                        
                    c_idx += 1
                r_idx += 1
                
        if not op_wb.sheetnames:
            op_wb.create_sheet(title="Sheet1")
        return op_wb

   
    # ==================== 合并单元格处理辅助工具 ====================
    def get_merged_cells_map(self, sheet):
        merged_map = {}
        for merged_range in sheet.merged_cells.ranges:
            top_left_val = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(r, c)] = top_left_val
        return merged_map

    def get_cell_value(self, sheet, row, col, merged_map):
        if (row, col) in merged_map:
            return merged_map[(row, col)]
        return sheet.cell(row=row, column=col).value
    # ==============================================================

    def show_software_intro(self):
        """打开单独的窗口，向用户展示软件的功能介绍和业务说明"""
        intro_win = tk.Toplevel(self)
        intro_win.title("软件说明与介绍")
        intro_win.geometry("660x550")
        intro_win.minsize(550, 450)
        intro_win.transient(self)  # 子窗口置于主窗口上层
        intro_win.grab_set()       # 开启模态拦截，关闭前无法操作主窗口

        # 文本展示区域
        txt_area = ScrolledText(intro_win, wrap=tk.WORD, font=("Microsoft YaHei", 10), spacing1=3, spacing3=3)
        txt_area.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)


        intro_text = """一、 软件基本定位与命名

  - 中文名： 流水预收快对
  - 英文名：QuickFlow
  - 定位：用于辅助财务人员进行“流水账目”与“预收款记录”自动核对与数据补全的本地工具。

二、 软件主要功能

1.  多格式账单自动读取与解析：
      - 支持直接读取和解析 .xlsx、旧版 .xls（包括系统导出的伪 XML 2003 xls 格式）以及多种编码的 .csv 账单文件。
      - 自动处理合并单元格，遇到多行共用一个日期或客户等情况时，会自动向下填充，防止漏读关键信息。

2.  财务信息自动匹配与补全（反写填充）：
      - 以“流水文件”为基准表，将“预收款记录”作为比对数据库。
      - 匹配成功后，软件会自动在流水文件后方的空白列中，追加写入对应的：单据编号、结算客户编码、客户名称、收款金额以及收款备注.
3.  多通道智能核对对账：
      - 银行流水：针对中文客户，结合交易金额、交易日期、币种以及清洗后的客户名称（自动去除空格和符号差异）进行多维度比对；针对英文客户，主要通过折合本位币的金额进行快速匹配。
      - 支付宝/微信：基于入账金额与入账日期（精确到天）进行快速匹配。

4.  双向“1对1”避错校验：
      - 遵循排他性比对原则。如果某笔流水在预收款中找到多笔同特征记录，或者多笔流水对应了同一笔预收款，软件会主动放弃自动关联并归为冲突项，交由人工核实，防止错误的资金关联。

5.  无损原始排版输出：
      - 输出结果时，不破坏原始流水文件原有的字体、颜色、边框等排版样式，不影响原本存在的计算公式和未变动列。

6.  未匹配原因诊断报告：
      - 运行完毕后，在同目录下自动生成一份 .txt 格式的诊断日志，清晰记录每一笔未能自动匹配成功的流水原因（如：金额不存在、交易日期不符、名称不符合、存在多笔重复特征等）。"""

        txt_area.insert(tk.END, intro_text)
        txt_area.configure(state='disabled') # 设置为只读

        # 关闭按钮
        close_btn = ttk.Button(intro_win, text="关闭说明", command=intro_win.destroy)
        close_btn.pack(pady=10)

    def start_processing_thread(self):
        if not self.bank_file_path.get() or not self.pr_file_path.get():
            messagebox.showwarning("警告", "请确保已指定流水文件与预收款文件！")
            return
        
        self.run_btn.configure(state=tk.DISABLED)
        self.status_text.set("正在核对数据中...")
        self.clear_log()
        
        t = threading.Thread(target=self.process_reconciliation)
        t.daemon = True
        t.start()

    def process_reconciliation(self):
        start_time = datetime.datetime.now()
        log_messages = []

        def local_log(msg):
            self.log(msg)
            log_messages.append(msg)

        mode = self.mode_var.get()
        cfg = MODE_CONFIGS[mode]

        bank_path = self.bank_file_path.get()
        pr_path = self.pr_file_path.get()
        
        # 提取银行流水所在目录及原名
        bank_dir, bank_full_filename = os.path.split(bank_path)
        bank_name_no_ext, _ = os.path.splitext(bank_full_filename)

        local_log(f"开始处理时间: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        local_log(f"当前对账模式: 【{mode}】")
        local_log(f"流水文件路径: {bank_path}")
        local_log(f"预收款路径: {pr_path}")
        local_log(f"结果输出目录: {bank_dir}")

        # 1. 读取并预解析预收款文件
        try:
            pr_wb = self.load_any_file_as_openpyxl(pr_path, data_only=True)
            pr_ws = self.get_sheet_insensitive(pr_wb, "sheet1")
            pr_merged_map = self.get_merged_cells_map(pr_ws)
        except Exception as e:
            local_log(f"错误: 读取预收款文件失败: {str(e)}")
            self.finalize_run(False, log_messages, bank_dir)
            return

        # 获取预收款表头
        pr_headers = [
            str(self.get_cell_value(pr_ws, PR_HEADER_ROW, col, pr_merged_map)).strip()
            if self.get_cell_value(pr_ws, PR_HEADER_ROW, col, pr_merged_map) is not None else ""
            for col in range(1, pr_ws.max_column + 1)
        ]
        
        # 匹配预收款各列索引
        col_pr_settle = self.find_header_col(pr_headers, [ "收款账户"])
        col_pr_amount = self.find_header_col(pr_headers, [ "收款金额"])
        col_pr_date = self.find_header_col(pr_headers, [ "单据日期"])
        col_pr_billno = self.find_header_col(pr_headers, [ "单据编号"])
        col_pr_client = self.find_header_col(pr_headers, [ "结算客户"])
        col_pr_client_code = self.find_header_col(pr_headers, ["结算客户编码"])
        col_pr_amount_local = self.find_header_col(pr_headers, [ "收款金额本位币"])
        col_pr_currency = self.find_header_col(pr_headers, [ "币别"])
        col_pr_comment = self.find_header_col(pr_headers, ["收款备注"])

        # 验证预收款中必需的列
        required_pr = {
            "收款账户": col_pr_settle,
            "收款金额": col_pr_amount,
            "单据日期": col_pr_date,
            "单据编号": col_pr_billno,
            "结算客户": col_pr_client,
            "结算客户编码": col_pr_client_code,
            "收款金额本位币": col_pr_amount_local
        }
        if mode == "银行":
            required_pr["币别"] = col_pr_currency
        elif mode in ("支付宝", "微信"):
            required_pr["收款备注"] = col_pr_comment

        missing_pr_cols = [k for k, v in required_pr.items() if v is None]
        if missing_pr_cols:
            local_log(f"错误：预收款文件中缺少匹配所需的列: {', '.join(missing_pr_cols)}")
            self.finalize_run(False, log_messages, bank_dir)
            return

        # 筛选加载预收款数据
        pr_items = []
        target_settle = cfg["pr_settle_account"]
        skipped_settle_count = 0

        for r in range(PR_DATA_START_ROW, pr_ws.max_row + 1):
            settle_val = self.get_cell_value(pr_ws, r, col_pr_settle, pr_merged_map)
            if settle_val is None:
                continue
            
            settle_str = str(settle_val).strip()
            if settle_str != target_settle:
                skipped_settle_count += 1
                continue

            amount_raw = self.get_cell_value(pr_ws, r, col_pr_amount, pr_merged_map)
            amount = self.clean_amount(amount_raw)
            if amount is None:
                local_log(f"提示 [预收款行号 {r}]: 收款金额为空，已跳过该行。")
                continue

            date_raw = self.get_cell_value(pr_ws, r, col_pr_date, pr_merged_map)
            date_val = self.parse_date_only(date_raw)
            if date_val is None:
                local_log(f"提示 [预收款行号 {r}]: 单据日期为空或格式错误，已跳过该行。")
                continue

            amount_local_raw = self.get_cell_value(pr_ws, r, col_pr_amount_local, pr_merged_map)
            amount_local = self.clean_amount(amount_local_raw)

            billno = str(self.get_cell_value(pr_ws, r, col_pr_billno, pr_merged_map) or "").strip()
            client_code = str(self.get_cell_value(pr_ws, r, col_pr_client_code, pr_merged_map) or "").strip()
            client = str(self.get_cell_value(pr_ws, r, col_pr_client, pr_merged_map) or "").strip()
            
            currency = ""
            if col_pr_currency:
                currency = str(self.get_cell_value(pr_ws, r, col_pr_currency, pr_merged_map) or "").strip()

            comment = ""
            if col_pr_comment:
                comment = str(self.get_cell_value(pr_ws, r, col_pr_comment, pr_merged_map) or "").strip()

            pr_items.append({
                "row_num": r,
                "amount": amount,
                "amount_local": amount_local,
                "date": date_val,
                "billno": billno,
                "client_code": client_code,
                "client": client,
                "currency": currency,
                "comment": comment
            })

        local_log(f"预收款过滤统计: 排除非【{target_settle}】数据记录共 {skipped_settle_count} 条。")
        local_log(f"预收款有效可用记录条数: {len(pr_items)}")

        # 2. 读取并预解析流水文件
        try:
            bank_wb_values = self.load_any_file_as_openpyxl(bank_path, data_only=True)
            bank_ws_values = self.get_sheet_insensitive(bank_wb_values, "sheet1")
            bank_merged_map = self.get_merged_cells_map(bank_ws_values)
            
            # 若原文件为 .xlsx 则完整载入非值模式保存原有排版与公式，若为 .xls 或 .csv 则直接使用当前已转换的工作簿另存为 .xlsx
            bank_ext = os.path.splitext(bank_path)[1].lower()
            if bank_ext == '.xlsx':
                bank_wb_save = openpyxl.load_workbook(bank_path, data_only=False)
            else:
                bank_wb_save = self.load_any_file_as_openpyxl(bank_path, data_only=True)
                
            bank_ws_save = self.get_sheet_insensitive(bank_wb_save, "sheet1")
        except Exception as e:
            local_log(f"错误: 读取流水文件失败: {str(e)}")
            self.finalize_run(False, log_messages, bank_dir)
            return

        # 获取流水表头
        bank_headers = [
            str(self.get_cell_value(bank_ws_values, cfg["bank_header_row"], col, bank_merged_map)).strip()
            if self.get_cell_value(bank_ws_values, cfg["bank_header_row"], col, bank_merged_map) is not None else ""
            for col in range(1, bank_ws_values.max_column + 1)
        ]
        
        # 匹配流水表所需定位字段
        bank_cols_idx = {}
        missing_bank_cols = []
        for key, field_name in cfg["bank_cols"].items():
            idx = None
            if key == "date" and mode == "微信":
                # 微信匹配包含问号的变种或普通形式
                idx = self.find_header_col(bank_headers, ["交易时间", "?交易时间", "？交易时间"])
            else:
                idx = self.find_header_col(bank_headers, [field_name])
            
            if idx is None:
                missing_bank_cols.append(field_name)
            else:
                bank_cols_idx[key] = idx

        if missing_bank_cols:
            local_log(f"错误：流水文件中未定位到必要字段列: {', '.join(missing_bank_cols)}")
            self.finalize_run(False, log_messages, bank_dir)
            return

        # 解析加载流水数据
        bank_rows_data = []
        for r in range(cfg["bank_data_start_row"], bank_ws_values.max_row + 1):
            row_vals = [self.get_cell_value(bank_ws_values, r, c, bank_merged_map) for c in range(1, bank_ws_values.max_column + 1)]
            if all(v is None for v in row_vals):
                continue

            amount_raw = self.get_cell_value(bank_ws_values, r, bank_cols_idx["amount"], bank_merged_map)
            amount = self.clean_amount(amount_raw)
            if amount is None:
                local_log(f"提示 [流水行号 {r}]: 金额字段未提取到有效数值，已忽略。")
                continue

            date_raw = self.get_cell_value(bank_ws_values, r, bank_cols_idx["date"], bank_merged_map)
            date_val = self.parse_date_only(date_raw)

            opp_name = ""
            if "opp_name" in bank_cols_idx:
                opp_name = str(self.get_cell_value(bank_ws_values, r, bank_cols_idx["opp_name"], bank_merged_map) or "").strip()

            currency = ""
            if "currency" in bank_cols_idx:
                currency = str(self.get_cell_value(bank_ws_values, r, bank_cols_idx["currency"], bank_merged_map) or "").strip()

            bank_rows_data.append({
                "row_num": r,
                "amount": amount,
                "date": date_val,
                "opp_name": opp_name,
                "currency": currency
            })

        local_log(f"流水文件有效提取到的流水行总计: {len(bank_rows_data)}")

        # 3. 对账匹配
        bank_to_pr = {}
        pr_to_bank = {}

        for br in bank_rows_data:
            br_num = br["row_num"]
            for pr in pr_items:
                pr_num = pr["row_num"]
                matched = False
                
                if mode == "银行":
                    is_chinese = self.is_chinese_name(br["opp_name"])
                    if is_chinese:
                        # 2.1 中文对方账户名称匹配
                        cond1 = (br["amount"] == pr["amount"])
                        cond2 = (br["date"] == pr["date"] and br["date"] is not None)
                        
                        # 清洗比较
                        norm_opp = self.clean_string_for_compare(br["opp_name"])
                        norm_client = self.clean_string_for_compare(pr["client"])
                        cond3 = (norm_opp == norm_client and norm_opp != "")
                        
                        norm_curr_bank = self.clean_string_for_compare(br["currency"])
                        norm_curr_pr = self.clean_string_for_compare(pr["currency"])
                        cond4 = (norm_curr_bank == norm_curr_pr and norm_curr_bank != "")
                        
                        if cond1 and cond2 and cond3 and cond4:
                            matched = True
                    else:
                        # 2.2 英文对方账户名称匹配
                        cond1 = (br["amount"] == pr["amount_local"] and pr["amount_local"] is not None)
                        if cond1:
                            matched = True
                            
                elif mode in ("支付宝", "微信"):
                    cond1 = (br["amount"] == pr["amount"])
                    cond2 = (br["date"] == pr["date"] and br["date"] is not None)
                    if cond1 and cond2:
                        matched = True

                if matched:
                    bank_to_pr.setdefault(br_num, []).append(pr)
                    pr_to_bank.setdefault(pr_num, []).append(br_num)

        # 4. 写入与统计
        matched_count = 0
        unmatched_count = 0
        duplicate_count = 0

        # 在表头行写入新增关联列标题
        for col_idx, (header_name, _) in cfg["output_cols"].items():
            bank_ws_save.cell(row=cfg["bank_header_row"], column=col_idx, value=header_name)

        local_log("\n" + "-"*15 + " 开始执行对账比对校验 " + "-"*15)

        for br in bank_rows_data:
            br_num = br["row_num"]
            br_amount = br["amount"]
            candidates = bank_to_pr.get(br_num, [])

            if len(candidates) == 1:
                pr_item = candidates[0]
                associated_bank_rows = pr_to_bank.get(pr_item["row_num"], [])
                
                # 双向确认，确保为唯一映射
                if len(associated_bank_rows) == 1:
                    for col_idx, (_, pr_field) in cfg["output_cols"].items():
                        val_to_write = pr_item.get(pr_field, "")
                        bank_ws_save.cell(row=br_num, column=col_idx, value=val_to_write)
                    matched_count += 1
                else:
                    unmatched_count += 1
                    duplicate_count += 1
                    local_log(
                        f"无法匹配 [多行关联 - 流水行号 {br_num}]: 流水行虽匹配到预收款行 {pr_item['row_num']} (金额: {br_amount})，"
                        f"但此预收款记录还对应了其他多个流水行号 {associated_bank_rows}，由于不具排他性，已执行跳过。"
                    )
            elif len(candidates) > 1:
                unmatched_count += 1
                duplicate_count += 1
                candidate_rows = [c["row_num"] for c in candidates]
                local_log(
                    f"无法匹配 [特征重复 - 流水行号 {br_num}]: 金额 {br_amount} 在预收款中匹配到多行同特征记录 (行号: {candidate_rows})，无法确认唯一，已执行跳过。"
                )
            else:
                unmatched_count += 1
                
                # 寻找同金额排除项，诊断过滤失败原因
                potential_prs = []
                for pr in pr_items:
                    is_amount_match = False
                    if mode == "银行":
                        is_chinese = self.is_chinese_name(br["opp_name"])
                        if is_chinese:
                            is_amount_match = (br["amount"] == pr["amount"])
                        else:
                            is_amount_match = (br["amount"] == pr["amount_local"])
                    else:
                        is_amount_match = (br["amount"] == pr["amount"])
                    
                    if is_amount_match:
                        potential_prs.append(pr)

                if potential_prs:
                    reasons_list = []
                    for p in potential_prs:
                        reasons = []
                        if mode == "银行":
                            is_chinese = self.is_chinese_name(br["opp_name"])
                            if is_chinese:
                                if br["date"] != p["date"]:
                                    reasons.append(f"日期不满足(流水:{br['date']} vs 预收:{p['date']})")
                                norm_opp = self.clean_string_for_compare(br["opp_name"])
                                norm_client = self.clean_string_for_compare(p["client"])
                                if norm_opp != norm_client:
                                    reasons.append(f"对方账户名不满足(流水:{br['opp_name']} vs 预收:{p['client']})")
                                norm_curr_bank = self.clean_string_for_compare(br["currency"])
                                norm_curr_pr = self.clean_string_for_compare(p["currency"])
                                if norm_curr_bank != norm_curr_pr:
                                    reasons.append(f"币别不满足(流水:{br['currency']} vs 预收:{p['currency']})")
                            else:
                                # 英文无其他强制必选规则
                                pass
                        else:
                            if br["date"] != p["date"]:
                                reasons.append(f"日期不满足(流水:{br['date']} vs 预收:{p['date']})")
                        
                        if reasons:
                            reasons_str = " 且 ".join(reasons)
                            reasons_list.append(f"预收款行 {p['row_num']}(具体: {reasons_str})")
                    
                    if reasons_list:
                        local_log(
                            f"未匹配 [条件过滤未通过 - 流水行号 {br_num}]: 金额 {br_amount} 的预收款存在，但未能通过匹配限制：\n"
                            f"   -> {'; '.join(reasons_list)}"
                        )
                    else:
                        local_log(
                            f"未匹配 [未找到符合记录 - 流水行号 {br_num}]: 预收款库中未能找到符合该行条件的可用记录。"
                        )
                else:
                    local_log(
                        f"未匹配 [无此金额记录 - 流水行号 {br_num}]: 未能在对应账户的预收款中检索到此金额 ({br_amount}) 的记录。"
                    )

        # 5. 输出保存
        end_time = datetime.datetime.now()
        timestamp_str = end_time.strftime("%Y%m%d_%H%M%S")
        
        # 结果文件名：原流水文件名 + 运行结束时间戳.xlsx (强制以 xlsx 保存)
        output_excel_path = os.path.join(bank_dir, f"{bank_name_no_ext}_{timestamp_str}.xlsx")
        output_log_path = os.path.join(bank_dir, f"{bank_name_no_ext}_{timestamp_str}.txt")
        
        try:
            bank_wb_save.save(output_excel_path)
            local_log(f"\n结果Excel文件已成功保存至: {output_excel_path}")
        except Exception as e:
            local_log(f"错误: 保存结果文件失败: {str(e)}")
            self.finalize_run(False, log_messages, bank_dir)
            return

        local_log("\n" + "="*15 + " 核对统计汇总 " + "="*15)
        local_log(f"完成处理时间: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        local_log(f"数据比对耗时: {(end_time - start_time).total_seconds():.2f} 秒")
        local_log(f"有效比对流水行总数: {len(bank_rows_data)}")
        local_log(f"完成唯一匹配条数: {matched_count}")
        local_log(f"未匹配/冲突条数: {unmatched_count}")
        local_log("="*44)

        try:
            with open(output_log_path, "w", encoding="utf-8") as f:
                for line in log_messages:
                    f.write(line + "\n")
            local_log(f"运行日志已输出并保存在: {output_log_path}")
        except Exception as e:
            local_log(f"警告: 保存外部日志文本文件失败: {str(e)}")

        self.finalize_run(True, log_messages, bank_dir)

    def finalize_run(self, success, log_messages, target_output_dir):
        self.run_btn.configure(state=tk.NORMAL)
        if success:
            self.status_text.set("核对完毕")
            messagebox.showinfo("成功", f"流水账目核对已执行完成！\n\n结果文件已保存在以下目录位置：\n\n{target_output_dir}")
        else:
            self.status_text.set("核对失败")
            messagebox.showerror("出错", "比对过程中遇到异常错误，详情请参照下方运行日志。")


if __name__ == "__main__":
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = ReconciliationApp()
    app.mainloop()